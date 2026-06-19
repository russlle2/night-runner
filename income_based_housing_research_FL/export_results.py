from __future__ import annotations

from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill

from config import CALL_SCRIPT_MD, FINAL_CSV, FINAL_JSON, FINAL_XLSX, README_MD, SOURCE_LOG_MD
from models import HousingProperty
from utils.evidence_utils import evidence_lines
from utils.io_utils import append_log, ensure_directories, read_json, write_json
from utils.scoring_utils import affordability_rank

EXPORT_COLUMNS = [
    "property_name",
    "property_type",
    "address",
    "city",
    "county",
    "state",
    "zip",
    "phone",
    "email",
    "website_url",
    "management_company",
    "management_company_phone",
    "housing_authority_or_agency",
    "source_urls",
    "source_quality",
    "bedroom_sizes",
    "unit_count",
    "population_served",
    "program_type",
    "rent_type",
    "exact_published_rent_by_bedroom",
    "exact_published_rent_by_income_bracket",
    "income_limits_by_household_size",
    "AMI_brackets_supported",
    "application_fee",
    "deposit",
    "utilities_included",
    "pet_policy",
    "criminal_background_policy",
    "credit_policy",
    "eviction_policy",
    "minimum_income_requirement",
    "maximum_income_requirement",
    "required_documents",
    "application_method",
    "application_url",
    "waitlist_status",
    "vacancy_status",
    "last_updated_date_from_source",
    "vacancy_likelihood_score",
    "vacancy_likelihood_reason",
    "phone_verification_needed",
    "notes",
    "confidence_score",
    "distance_miles_to_target_area",
]


def load_properties() -> list[HousingProperty]:
    payload = read_json(Path("intermediate") / "properties_final.json", default=[])
    return [HousingProperty(**item) for item in payload]


def flatten_properties(properties: list[HousingProperty]) -> list[dict]:
    rows = []
    for item in properties:
        row = item.export_dict()
        rows.append({column: row.get(column, "") for column in EXPORT_COLUMNS})
    return rows


def write_source_log(properties: list[HousingProperty]) -> None:
    lines = ["# Source Log", ""]
    for property_record in properties:
        lines.append(f"## {property_record.property_name}")
        lines.append(f"- address: {property_record.address}")
        lines.append(f"- phone: {property_record.phone or 'NOT FOUND / NEEDS CALL'}")
        lines.append(f"- source_quality: {property_record.source_quality}")
        lines.append("- confirmed evidence:")
        lines.extend(evidence_lines(property_record) or ["- none captured"])
        lines.append(f"- unverified items: vacancy={property_record.vacancy_status}, waitlist={property_record.waitlist_status}, rent={property_record.exact_published_rent_by_bedroom}")
        lines.append(f"- phone call needed: {'yes' if property_record.phone_verification_needed else 'no'}")
        lines.append("")
    SOURCE_LOG_MD.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def write_call_script(properties: list[HousingProperty]) -> None:
    lines = [
        "# Call Script",
        "",
        "Hi, I’m looking for income-based or affordable housing. Are you currently accepting applications or is your waitlist open? What bedroom sizes do you have? What is the rent based on? Do you have any units available now or expected soon? What income limits apply? What documents do I need to apply? Is there an application fee or deposit? Can I apply online or do I need to come in person?",
        "",
        "## Call Tracking Table",
        "",
        "| property_name | phone | date_called | person_spoken_to | vacancy_status | waitlist_status | rent_details | documents_needed | next_step | notes |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for property_record in sorted(properties, key=lambda item: affordability_rank(item.model_dump())):
        lines.append(
            f"| {property_record.property_name} | {property_record.phone or 'NOT FOUND / NEEDS CALL'} |  |  | "
            f"{property_record.vacancy_status} | {property_record.waitlist_status} | "
            f"{property_record.exact_published_rent_by_bedroom} | {property_record.required_documents} |  | {property_record.vacancy_likelihood_reason} |"
        )
    CALL_SCRIPT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def summary_sections(df: pd.DataFrame) -> str:
    top_ranked = df.sort_values(
        by=["vacancy_likelihood_score", "confidence_score"], ascending=[False, False]
    ).head(10)[["property_name", "city", "program_type", "phone", "vacancy_likelihood_score", "confidence_score"]]
    truly_income_based = df[df["rent_type"].isin(["income-based", "max LIHTC rent"])]
    exact_rent = df[~df["exact_published_rent_by_bedroom"].str.contains("NOT FOUND", na=False)]
    rent_verify = df[df["exact_published_rent_by_bedroom"].str.contains("NOT FOUND", na=False)]
    senior_disabled = df[df["property_type"].str.contains("Senior|Disabled", case=False, na=False)]
    family_general = df[df["population_served"].str.contains("family|general", case=False, na=False)]
    common_docs = Counter()
    for docs in df["required_documents"].fillna(""):
        for part in [item.strip() for item in docs.split(";") if item.strip() and "NOT FOUND" not in item]:
            common_docs[part] += 1

    warnings = []
    if not df[df["waitlist_status"] == "closed"].empty:
        warnings.append("Some properties explicitly appear to have closed waitlists.")
    if not df[df["vacancy_status"] == "call_required"].empty:
        warnings.append("Many properties still require phone verification for current availability.")
    if df[df["source_quality"].isin(["apartment_listing", "nonprofit_directory", "unknown"])].shape[0]:
        warnings.append("Several records depend partly on third-party listings and should be verified against the property directly.")

    def frame_to_markdown(frame: pd.DataFrame) -> str:
        if frame.empty:
            return "No properties found."
        headers = list(frame.columns)
        lines = [
            "| " + " | ".join(headers) + " |",
            "| " + " | ".join(["---"] * len(headers)) + " |",
        ]
        for _, row in frame.iterrows():
            values = [str(row[column]).replace("\n", " ").strip() for column in headers]
            lines.append("| " + " | ".join(values) + " |")
        return "\n".join(lines)

    return "\n".join(
        [
            "## Plain-English Summary",
            "",
            "### Best options to call first",
            frame_to_markdown(top_ranked),
            "",
            "### Properties most likely to have vacancy",
            frame_to_markdown(
                df.sort_values(["vacancy_likelihood_score", "confidence_score"], ascending=[False, False])[
                    ["property_name", "city", "vacancy_likelihood_score", "vacancy_likelihood_reason"]
                ].head(10)
            ),
            "",
            "### Properties most likely to be truly income-based",
            frame_to_markdown(truly_income_based[["property_name", "city", "program_type", "rent_type"]].head(10))
            if not truly_income_based.empty
            else "No strongly confirmed income-based records were found.",
            "",
            "### Properties with exact rent published",
            frame_to_markdown(exact_rent[["property_name", "city", "exact_published_rent_by_bedroom"]].head(10))
            if not exact_rent.empty
            else "No exact published rents were captured.",
            "",
            "### Properties where rent requires phone/application verification",
            frame_to_markdown(rent_verify[["property_name", "city", "rent_type"]].head(10))
            if not rent_verify.empty
            else "All properties had rent figures published.",
            "",
            "### Properties with senior/disabled restrictions",
            frame_to_markdown(senior_disabled[["property_name", "city", "property_type"]].head(10))
            if not senior_disabled.empty
            else "No senior/disabled restrictions confirmed.",
            "",
            "### Properties with family/general eligibility",
            frame_to_markdown(family_general[["property_name", "city", "population_served"]].head(10))
            if not family_general.empty
            else "No family/general eligibility notes confirmed.",
            "",
            "### Application documents most commonly required",
            "\n".join([f"- {name}: {count}" for name, count in common_docs.most_common(10)]) or "- No common document list was confirmed online.",
            "",
            "### Important warnings",
            "\n".join([f"- {warning}" for warning in warnings]) or "- No major warnings beyond normal property verification needs.",
            "",
        ]
    )


def write_readme(df: pd.DataFrame) -> None:
    readme = "\n".join(
        [
            "# Florida Income-Based Housing Research Workflow",
            "",
            "## Installation",
            "",
            "```bash",
            "python3 -m venv .venv",
            "source .venv/bin/activate",
            "pip install -r requirements.txt",
            "python -m playwright install chromium",
            "```",
            "",
            "## Run the scraper",
            "",
            "```bash",
            "python run_all.py",
            "```",
            "",
            "## Rerun only one city",
            "",
            "```bash",
            "python search_sources.py --city \"Leesburg\"",
            "python scrape_pages.py --city \"Leesburg\"",
            "python normalize_properties.py",
            "python enrich_official_sources.py",
            "python rent_extractor.py",
            "python apply_requirements_extractor.py",
            "python vacancy_estimator.py",
            "python export_results.py",
            "```",
            "",
            "## Open the final Excel file",
            "",
            "Open `housing_results.xlsx` in Excel, LibreOffice Calc, or Numbers.",
            "",
            "## Manual verification guidance",
            "",
            "- Start with the `Call Required` sheet in the Excel workbook.",
            "- Use `call_script.md` to track phone calls.",
            "- Prioritize records with higher vacancy scores and stronger confidence scores.",
            "- Treat any `NOT FOUND` or `NEEDS CALL` fields as unresolved until confirmed directly by the property.",
            "",
            summary_sections(df),
        ]
    )
    README_MD.write_text(readme + "\n", encoding="utf-8")


def format_excel() -> None:
    workbook = load_workbook(FINAL_XLSX)
    vacancy_fills = {
        4: PatternFill("solid", fgColor="70AD47"),
        3: PatternFill("solid", fgColor="C6E0B4"),
        2: PatternFill("solid", fgColor="FFD966"),
        1: PatternFill("solid", fgColor="F4B183"),
        0: PatternFill("solid", fgColor="E06666"),
    }

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        if "vacancy_likelihood_score" in [cell.value for cell in sheet[1]]:
            header_map = {cell.value: cell.column_letter for cell in sheet[1]}
            column = header_map["vacancy_likelihood_score"]
            for score, fill in vacancy_fills.items():
                sheet.conditional_formatting.add(
                    f"{column}2:{column}{sheet.max_row}",
                    CellIsRule(operator="equal", formula=[str(score)], fill=fill),
                )

    workbook.save(FINAL_XLSX)


def main() -> None:
    ensure_directories()
    properties = load_properties()
    properties = sorted(properties, key=lambda item: (-item.vacancy_likelihood_score, -item.confidence_score, item.property_name))
    rows = flatten_properties(properties)
    df = pd.DataFrame(rows, columns=EXPORT_COLUMNS)
    df.to_csv(FINAL_CSV, index=False)
    write_json(FINAL_JSON, [item.model_dump() for item in properties])

    with pd.ExcelWriter(FINAL_XLSX, engine="openpyxl") as writer:
        df.sort_values(["vacancy_likelihood_score", "confidence_score"], ascending=[False, False]).to_excel(
            writer, index=False, sheet_name="All Properties"
        )
        df[df["vacancy_likelihood_score"] >= 3].sort_values(
            ["vacancy_likelihood_score", "confidence_score"], ascending=[False, False]
        ).to_excel(writer, index=False, sheet_name="Highest Vacancy Likelihood")
        df[df["rent_type"].isin(["income-based", "max LIHTC rent", "fixed affordable rent"])].to_excel(
            writer, index=False, sheet_name="Lowest Rent - Income-Based"
        )
        df[df["phone_verification_needed"] == True].to_excel(writer, index=False, sheet_name="Call Required")  # noqa: E712
        df[df["source_quality"].isin(["official_property_site", "government", "housing_authority", "HUD/USDA/FHFC"])].to_excel(
            writer, index=False, sheet_name="Official Sources"
        )
        df[["property_name", "phone", "required_documents", "application_fee", "deposit", "application_method", "application_url"]].to_excel(
            writer, index=False, sheet_name="Application Checklist"
        )
        source_log_rows = [
            {
                "property_name": item.property_name,
                "source_urls": "; ".join(item.source_urls),
                "vacancy_reason": item.vacancy_likelihood_reason,
                "notes": item.notes,
            }
            for item in properties
        ]
        pd.DataFrame(source_log_rows).to_excel(writer, index=False, sheet_name="Source Log")

    format_excel()
    write_source_log(properties)
    write_call_script(properties)
    write_readme(df)
    append_log(f"export_results.py: exported {len(properties)} properties")


if __name__ == "__main__":
    main()
